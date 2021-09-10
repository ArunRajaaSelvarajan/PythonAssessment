import os
from shutil import which

import openpyxl
import pytest
import pandas as pd
from selenium import webdriver

from pageObjects.CandidatesPage import CandidatesPage
from pageObjects.LandingPage import LandingPage
from pageObjects.LoginPage import LoginPage
from utilities.BaseClass import BaseClass


class TestSuite(BaseClass):

    def test_validate_data(self):
        # Object initiation
        log = self.getLogger()
        login = LoginPage(self.driver)

        # Enter credentials and click submit
        log.info("Entering login details and click submit button")
        login.username().send_keys("Admin")
        login.password().send_keys("admin123")
        login.submit().click()

        # Get value of content text
        landingPage = LandingPage(self.driver)
        contentText = landingPage.content().text

        log.info("Get the name of content and validate whether it is equal to 'Dashboard'")
        if contentText == "Dashboard":
            log.info("Dashboard tab is selected as expected")

        # select 'Recruitment' tab
        log.info("Select 'Recruitment' tab")
        self.driver.find_element_by_css_selector("a[id='menu_recruitment_viewRecruitmentModule']").click()

        # select 'Candidates' tab
        log.info("Select 'Candidates' sub tab")
        landingPage.candidate().click()

        if contentText == "Candidates":
            log.info("'Candidates' tab is selected by default")

        candidate = CandidatesPage(self.driver)
        tableRows = candidate.tableRow()
        rowCount = len(tableRows)

        lstVacancy = []
        lstCandidate = []
        lstHiringManager = []
        lstDOA = []
        lstStatus = []

        # Code to get all the values from data table and store it in dataframe
        log.info("Get all the values from data table and store it in dataframe")

        for count in range(1, len(tableRows) + 1):
            textVacancy = self.driver.find_element_by_xpath(
                "//table[@id='resultTable']/tbody/tr[" + str(count) + "]/td[2]").text
            textCand = self.driver.find_element_by_xpath(
                "//table[@id='resultTable']/tbody/tr[" + str(count) + "]/td[3]").text
            textHM = self.driver.find_element_by_xpath(
                "//table[@id='resultTable']/tbody/tr[" + str(count) + "]/td[4]").text
            textDOA = self.driver.find_element_by_xpath(
                "//table[@id='resultTable']/tbody/tr[" + str(count) + "]/td[5]").text
            textStatus = self.driver.find_element_by_xpath(
                "//table[@id='resultTable']/tbody/tr[" + str(count) + "]/td[6]").text

            lstVacancy.append(textVacancy)
            lstCandidate.append(textCand)
            lstHiringManager.append(textHM)
            lstDOA.append(textDOA)
            lstStatus.append(textStatus)

        dataFrame = pd.DataFrame()
        dataFrame['Vacancy'] = lstVacancy
        dataFrame['Candidate'] = lstCandidate
        dataFrame['Hiring Manager'] = lstHiringManager
        dataFrame['Date of Application'] = lstDOA
        dataFrame['Status'] = lstStatus
        dataFrame['Date of Application'] = dataFrame['Date of Application'].astype(str)

        print("Dataframe Contents fetched from Application:")
        print(dataFrame.to_string())

        # Code to read the data from excel and store in dataframe
        book = openpyxl.load_workbook(os.getcwd() + "/Input data.xlsx")
        sheet = book.active

        lstVacancy1 = []
        lstCandidate1 = []
        lstHiringManager1 = []
        lstDOA1 = []
        lstStatus1 = []

        for i in range(2, sheet.max_row + 1):
            textVacancy1 = sheet.cell(row=i, column=1).value
            textCand1 = sheet.cell(row=i, column=2).value
            textHM1 = sheet.cell(row=i, column=3).value
            textDOA1 = sheet.cell(row=i, column=5).value
            textStatus1 = sheet.cell(row=i, column=6).value

            lstVacancy1.append(textVacancy1)
            lstCandidate1.append(textCand1)
            lstHiringManager1.append(textHM1)
            lstDOA1.append(textDOA1)
            lstStatus1.append(textStatus1)

        dataFrame1 = pd.DataFrame()
        dataFrame1['Vacancy'] = lstVacancy1
        dataFrame1['Candidate'] = lstCandidate1
        dataFrame1['Hiring Manager'] = lstHiringManager1
        dataFrame1['Date of Application'] = lstDOA1
        dataFrame1['Status'] = lstStatus1
        dataFrame1['Date of Application'] = dataFrame1['Date of Application'].astype(str)

        print("Dataframe Contents fetched from excel:")
        print(dataFrame1.to_string())

        # compare dataframes and print the entries from excel that does not match with application
        df = pd.DataFrame()
        df = dataFrame1.merge(dataFrame, how='outer', indicator=True).loc[lambda x: x['_merge'] == 'left_only']

        print("Excel data entries not present in the application are as follows :")
        print(df.to_string())